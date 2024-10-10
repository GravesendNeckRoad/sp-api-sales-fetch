# _____________________________________________________________________________________________________________________
#                                               FOR INTERNAL USE ONLY
#                                               [REDACTED] TRADING CO.
# _____________________________________________________________________________________________________________________

import random
import time
from typing import List, Union
import openpyxl as xl


class Style:
    """
    Simple styling tool to format an .xlsx worksheet with the openpyxl library.

    :param ws: (xl.worksheet.worksheet.Worksheet) The worksheet of the open workbook you wish to format

    Example:
    import openpyxl as xl
    wb = xl.load_workbook(path)
    ws = wb.active
    styler = Style(ws)
    """
    def __init__(self, ws) -> None:
        self.ws = ws

    def apply_styles_to_cell(self, cell: str, bold: bool = True, highlighter: bool = True, color: str = None) -> None:
        """
        Applies style formatting to a desired cell in the openpyxl worksheet.

        :param cell: (str) The cell you wish to format. (Ex: 'C2')
        :param bold: (bool) If true, the cell will be made bold
        :param highlighter: (bool) If true, the cell will be highlighted
        :param color: (str) The color to highlight your column. Default = yellow
        """
        if not color:
            color = 'FFFF00'

        if highlighter:
            self.ws[cell].fill = xl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
        if bold:
            self.ws[cell].font = xl.styles.Font(bold=True)

    def currency_formatter(self, column_list: Union[List[str], str], max_row=None, currency: bool = True) -> None:
        """
        Formats every cell in a specified range as currency, or simply as a number with thousands separator.

        :param column_list: (str) The column letters you want to highlight (ex: ['K', 'L'], or 'E')
        :param max_row: (int) Specify a row number you want it to stop at. Default formats the entire column.
        :param currency: (bool) if False, then it will simply return the column formatted with a thousands separator.
        """
        if isinstance(column_list, str):
            column_list = list(column_list)

        if max_row is None:
            max_row = self.ws.max_row

        for col in column_list:
            for row in range(2, max_row + 1):
                if currency:
                    self.ws[f"{col}{row}"].number_format = '$#,##0.00'
                elif not currency:
                    self.ws[f"{col}{row}"].number_format = '#,##0'

    def align_and_center(self, start_row: int = 1, padding: int = 5) -> None:
        """ Auto align and widen all columns of your worksheet.

        :param start_row: (int) If your headers are long and/or text-wrapped, use >=2 to exclude headers as a reference.
        :param padding: (int) Add or remove whitespace from the columns
        """
        for row in self.ws.iter_cols(min_row=start_row):
            column_letter = xl.utils.get_column_letter(row[0].column)
            max_length = 0
            # first, widen each cell
            for cell in row:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                self.ws.column_dimensions[column_letter].width = max_length + padding
                # then, center align
                cell.alignment = xl.styles.Alignment(
                    horizontal='center',
                    vertical='center'
                )

    def create_table(self) -> None:
        """Formats an Excel array as a table, by identifying the first/last rows and columns of the worksheet."""
        last_column = xl.utils.get_column_letter(len(list(self.ws.columns)))
        last_row = self.ws.max_row
        table_range = f"A1:{last_column}{last_row}"

        create_table = xl.worksheet.table.Table(
            displayName=f"Table1",
            ref=table_range
        )

        table_design = xl.worksheet.table.TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )

        create_table.tableStyleInfo = table_design
        self.ws.add_table(create_table)


class Helpers:
    def __init__(self):
        pass

    @staticmethod
    def exponential_backoff(n, rate_of_growth=1.5, base_seconds=2, jitter=.01) -> None:
        """Simple timer function to manage API throttling."""
        x = (base_seconds * (rate_of_growth ** n))
        y = (random.uniform(-jitter*x, jitter*x))
        print(f"\tRetry attempt #{n} - {x+y:.2f} seconds ...")
        time.sleep(x+y)
